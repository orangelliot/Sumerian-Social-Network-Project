# Nicholas J Uhlhorn
# February 2022
# A program to translate the data from tabulatedData.xlsx into a
# dictionary with names as keys and a list of tablet ids as values
# This dictionary can be exported and also loaded in

# Usage: python3 nameDictionary.py data_path dictionary_output [imported_dictionary_path]
#       data_path - the data containing the tablets and the names on them
#       dictionary_output - the place to save the dictionary
#       imported_dictionary_path - an optional argument to start with an existing dictionary 

import sys
import os
from openpyxl import load_workbook, Workbook

# Check and get arguments
# Check number of args
if len(sys.argv) <= 2 or len(sys.argv) > 4:
    print("Incorrect number of arguments, see useage:")
    print("python3 nameDictionary.py data_path dictionary_output [imported_dictionary_path]")
    exit(0)
# Get input file
data_path = sys.argv[1]
dictionary_output = sys.argv[2]
if len(sys.argv) == 4:
    imported_dictionary_path = sys.argv[3]

# try to read the paths
try:
    data = load_workbook(data_path)
except:
    print("Problem with opening file: " + data_path)
    exit(0)

if len(sys.argv) == 4:
    try:
        imported_dictionary = load_workbook(imported_dictionary_path)
    except:
        print("Problem with opening file: " + imported_dictionary_path)
        exit(0)

dictionary = {}

# TODO: read imported dictionary into dictionary

# read data into the dictionary
data_sheet = data.worksheets[0]
for row in data_sheet.values:
    # get number of names on tablet
    try:
        name_count = int(row[1])
    except:
        continue # this skips the first row
    # iterate names in dictionary
    for name in row[2:2+name_count]:
        tablet_id = row[0][:7]
        if name not in dictionary.keys():
            # add new key
            dictionary[name] = []
        # add tablet to name entry if it isn't there already
        if tablet_id not in dictionary[name]:
            dictionary[name].append(tablet_id)

# output data to file
output = Workbook()
output.create_sheet('dictionary')

output_sheet = output.worksheets[0]

current_row = 2

for key in dictionary.keys():
    output_sheet.cell(current_row, 1, key)
    # set the current column to 1
    current_column = 2
    for tablet in dictionary[key]:
        # place name in cell and iterate current_column
        output_sheet.cell(current_row, current_column, tablet)
        current_column += 1
    current_row += 1

# save sheet
output.save(sys.argv[2])