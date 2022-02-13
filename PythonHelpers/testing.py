import openpyxl
import os

print('\n' + os.getcwd() + '\n')
print(os.listdir())

wb = openpyxl.load_workbook(filename= 'test.xlsx')
ws = wb.worksheets[0]

print(ws['A1'].value)

ws['A2'].value = 'beep'
ws.cell(2,2).value = 'boop'

print(ws['A2'].value)

wb.save('test.xlsx')