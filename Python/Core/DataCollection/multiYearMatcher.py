import os
import multiprocessing as mp
import psutil
from Database.SQLfuncs import SQLfuncs

import openpyxl
from difflib import SequenceMatcher

@staticmethod
def get_sim_metric(s1, s2):
    return SequenceMatcher(None, s1, s2).ratio()

def match_year(year_name):
    best_year = ''
    row = 1
    sim_metric = 0
    cur_year = cdli_years.cell(row, 1).value
    while cur_year != None:
        temp = get_sim_metric(cur_year, year_name)
        if  temp >= sim_metric:
            sim_metric = temp
            best_year = cdli_years.cell(row, 3).value
        row += 1
        cur_year = cdli_years.cell(row, 1).value
    return best_year, sim_metric

def thread_function(years, cpu, progress):
    progress[cpu - 1] = 0
    db = SQLfuncs('sumerian-social-network.clzdkdgg3zul.us-west-2.rds.amazonaws.com', 'root', '2b928S#%')
    for tuple in years:
        progress[cpu - 1] += 1
        row = 0
        best_sim = 0
        best_year = 'start'
        year = tuple[0]
        tablet = tuple[1]
        for i in range(cdli_years.cell(1,5).value):
            temp_year, similarity = match_year(year)
            if similarity >= best_sim:
                best_year = temp_year
                best_sim = similarity
        best_sim *= 100.0
        db.addBestYearToTab(best_year, tablet, str(int(best_sim)))
        row += 1

catalog = openpyxl.load_workbook(filename = 'catalog.xlsx')
cdli_years = catalog.worksheets[0]

if __name__ == '__main__':
    db = SQLfuncs('sumerian-social-network.clzdkdgg3zul.us-west-2.rds.amazonaws.com', 'root', '2b928S#%')
    n_cpus = psutil.cpu_count()
    procs = list()
    progress = mp.Array('i', range(n_cpus))
    years = db.execute_query("select * from rawyears group by tabid;")
    print(len(years))
    num_years = len(years)
    thread_size = int(num_years/n_cpus)
    pos = 0
    for cpu in range(n_cpus - 1):
        proc = mp.Process(target=thread_function, args=(years[pos:(pos + thread_size - 1)], cpu, progress,))
        procs.append(proc)
        pos += thread_size
    proc = mp.Process(target=thread_function, args=(years[pos:(num_years - 1)], n_cpus, progress,))
    procs.append(proc)

    for p in procs:
        p.start()

    sum = 0
    while sum < num_years:
        sum = 0
        for i in range(n_cpus):
            sum += progress[i]
        print("%d/%d" % (sum, num_years), end='\r')

    for p in procs:
        p.join()